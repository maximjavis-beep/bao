"""FBA 原始装箱数据解析器（openpyxl 版，无 pandas 依赖）"""
from pathlib import Path
import openpyxl


class FBAParser:
    COLUMN_MAP = {
        "msku": ["MSKU"],
        "asin": ["ASIN"],
        "title": ["标题", "品名"],
        "input_hs_code": ["进口海关编码"],
        "declared_qty": ["商品申报量"],
        "boxed_qty": ["商品装箱量"],
        "box_count": ["箱数"],
        "length": ["箱子长（cm）"],
        "width": ["箱子宽（cm）"],
        "height": ["箱子高（cm）"],
        "weight": ["箱子重量（kg）"],
        "box_range": ["货件箱子编号"],
    }

    def parse(self, file_path: str) -> dict:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb["packing"] if "packing" in wb.sheetnames else wb.active
        meta = self._extract_meta(ws)
        header_row = self._find_header_row(ws)
        if header_row < 0:
            wb.close()
            return {"meta": meta, "items": []}
        col_map = self._build_column_map(ws, header_row)
        items = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            msku = self._cell(ws, row_idx, col_map.get("msku"))
            if not msku or msku.strip() == "":
                continue
            items.append({
                "msku": msku.strip(),
                "title": self._cell(ws, row_idx, col_map.get("title")),
                "input_hs_code": self._cell(ws, row_idx, col_map.get("input_hs_code")),
                "asin": self._cell(ws, row_idx, col_map.get("asin")),
                "declared_qty": self._as_float(ws, row_idx, col_map.get("declared_qty")),
                "boxed_qty": self._as_float(ws, row_idx, col_map.get("boxed_qty")),
                "box_count": self._as_int(ws, row_idx, col_map.get("box_count")),
                "length": self._as_float(ws, row_idx, col_map.get("length")),
                "width": self._as_float(ws, row_idx, col_map.get("width")),
                "height": self._as_float(ws, row_idx, col_map.get("height")),
                "weight": self._as_float(ws, row_idx, col_map.get("weight")),
                "box_range": self._cell(ws, row_idx, col_map.get("box_range")),
            })
        wb.close()
        return {"meta": meta, "items": items}

    def _extract_meta(self, ws) -> dict:
        meta = {}
        for row_idx in range(1, min(9, ws.max_row + 1)):
            label = self._cell(ws, row_idx, 0)
            value = self._cell(ws, row_idx, 1)
            if not label:
                continue
            lbl = label.strip()
            if "货件编号" in lbl: meta["shipment_id"] = value.strip()
            elif "箱子数量" in lbl: meta["total_boxes"] = self._as_int(ws, row_idx, 1)
            elif "SKU 数量" in lbl or "SKU数量" in lbl: meta["sku_count"] = self._as_int(ws, row_idx, 1)
            elif "商品数量" in lbl: meta["item_count"] = self._as_int(ws, row_idx, 1)
        return meta

    def _find_header_row(self, ws) -> int:
        for row_idx in range(1, min(21, ws.max_row + 1)):
            for col_idx in range(min(20, ws.max_column or 0)):
                val = self._cell(ws, row_idx, col_idx)
                if val and "MSKU" in val:
                    return row_idx
        return -1

    def _build_column_map(self, ws, header_row: int) -> dict:
        col_map = {}
        for col_idx in range(min(30, ws.max_column or 0)):
            val = self._cell(ws, header_row, col_idx)
            if not val:
                continue
            for field, keywords in self.COLUMN_MAP.items():
                for kw in keywords:
                    if kw in val:
                        col_map[field] = col_idx
                        break
        return col_map

    @staticmethod
    def _cell(ws, row: int, col) -> str:
        if col is None or col < 0:
            return ""
        try:
            val = ws.cell(row=row + 1, column=col + 1).value
            return "" if val is None else str(val)
        except Exception:
            return ""

    @staticmethod
    def _as_float(ws, row: int, col) -> float:
        val = FBAParser._cell(ws, row, col)
        if not val:
            return 0.0
        try:
            return float(val)
        except Exception:
            return 0.0

    @staticmethod
    def _as_int(ws, row: int, col) -> int:
        val = FBAParser._cell(ws, row, col)
        if not val:
            return 0
        try:
            return int(float(val))
        except Exception:
            return 0
