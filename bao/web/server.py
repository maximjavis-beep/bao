"""Web 面板 — FBA 装箱单生成 HTTP 服务"""
import base64
import io
import json
import os
import shutil
import tempfile
import uuid
import zipfile
from datetime import date
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from urllib.parse import unquote

import openpyxl
from openpyxl.styles import PatternFill

from ..core.exporter import FBAExporter
from ..core.weaver import weave_fba
from ..parsers.fba_parser import FBAParser

WEB_DIR = Path(__file__).parent
DOWNLOADS_DIR = WEB_DIR / "downloads"
UPLOAD_DIR = Path(tempfile.gettempdir()) / "bao_uploads"


class BaoHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        p = self.path.split("?")[0]
        if p in ("/", "/index.html"):
            self._send_file("index.html", "text/html; charset=utf-8")
        elif p.startswith("/downloads/"):
            fp = DOWNLOADS_DIR / Path(unquote(p)).name
            if fp.exists():
                data = fp.read_bytes()
                self.send_response(200)
                self.send_header(
                    "Content-Type",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
            else:
                self.send_error(404)
        elif p == "/api/download-zip":
            self._handle_download_zip()
        else:
            self.send_error(404)

    def do_POST(self):
        p = self.path.split("?")[0]
        if p == "/api/parse":
            self._handle_parse()
        elif p == "/api/weave":
            self._handle_weave()
        elif p == "/api/weave-batch":
            self._handle_weave_batch()
        elif p == "/api/merge-plan":
            self._handle_merge_plan()
        else:
            self.send_error(404)

    def _read(self):
        return self.rfile.read(int(self.headers.get("Content-Length", 0)))

    def _json(self, data, status=200):
        b = json.dumps(data, ensure_ascii=False, default=str).encode()
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Content-Length", str(len(b)))
        self.end_headers()
        self.wfile.write(b)

    def _send_file(self, path, ct="application/octet-stream"):
        fp = WEB_DIR / path
        if not fp.exists():
            self.send_error(404)
            return
        data = fp.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", ct)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def _save(self, b64, prefix):
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        fp = UPLOAD_DIR / f"{prefix}_{uuid.uuid4().hex[:6]}.xlsx"
        fp.write_bytes(base64.b64decode(b64))
        return str(fp)

    @staticmethod
    def _parse_tracking(path):
        """解析货件追踪码文件 → {FBA编号: {tracking_code, warehouse, channel, total_boxes}}"""
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        # 按表头关键词自动定位列
        col_map = {}
        for c in range(1, min(20, ws.max_column or 0) + 1):
            h = str(ws.cell(row=1, column=c).value or "").strip()
            hl = h.lower()
            if "fba编号" in hl or "fba" in hl:
                col_map["fba"] = c
            elif "货件追踪" in h or "追踪码" in h or "tracking" in hl:
                col_map["tracking"] = c
            elif "仓库代码" in h or "仓库" in h:
                col_map["warehouse"] = c
            elif "渠道" in h:
                col_map["channel"] = c
            elif "箱数" in h or "总件数" in h:
                col_map["boxes"] = c
        # 兜底：旧格式 列1=FBA编号, 列2=追踪码
        if "fba" not in col_map:
            col_map["fba"] = 1
        if "tracking" not in col_map:
            col_map["tracking"] = 2
        result = {}
        for r in range(2, ws.max_row + 1):
            fba_id = str(ws.cell(row=r, column=col_map["fba"]).value or "").strip()
            if not fba_id:
                continue
            info = {"tracking_code": str(ws.cell(row=r, column=col_map["tracking"]).value or "").strip()}
            if "warehouse" in col_map:
                info["warehouse"] = str(ws.cell(row=r, column=col_map["warehouse"]).value or "").strip()
            if "channel" in col_map:
                info["channel"] = str(ws.cell(row=r, column=col_map["channel"]).value or "").strip()
            if "boxes" in col_map:
                v = ws.cell(row=r, column=col_map["boxes"]).value
                if v is not None:
                    info["total_boxes"] = int(v) if isinstance(v, (int, float)) else str(v).strip()
            result[fba_id] = info
        wb.close()
        return result

    def _handle_parse(self):
        """解析上传的 FBA 货件文件，返回预览数据"""
        try:
            body = json.loads(self._read())
            parser = FBAParser()

            b64 = body.get("file", "")
            if not b64:
                self._json({"success": False, "error": "未提供文件"}, 400)
                return

            path = self._save(b64, "fba")
            data = parser.parse(path)
            meta = data.get("meta", {})
            items = data.get("items", [])

            self._json({
                "success": True,
                "meta": meta,
                "items": items,
                "item_count": len(items),
                "path": path,
            })
        except Exception as e:
            self._json({"success": False, "error": str(e)}, 400)

    def _handle_weave(self):
        """编织生成 FBA 装箱单并返回下载链接"""
        try:
            body = json.loads(self._read())
            parser = FBAParser()

            b64 = body.get("file", "")
            hs_code = body.get("hs_code", None)
            tpl_b64 = body.get("template", None)
            track_b64 = body.get("tracking", None)

            if not b64:
                self._json({"success": False, "error": "未提供文件"}, 400)
                return

            path = self._save(b64, "fba")
            tpl_path = None
            if tpl_b64:
                tpl_path = self._save(tpl_b64, "tpl")
            tracking_map = None
            if track_b64:
                tracking_map = self._parse_tracking(self._save(track_b64, "track"))

            data = parser.parse(path)
            meta = data.get("meta", {})
            sid = meta.get("shipment_id", "UNKNOWN")

            woven = weave_fba(data, hs_code_override=hs_code)

            fname = f"{sid}-装箱单_{uuid.uuid4().hex[:6]}.xlsx"
            out = UPLOAD_DIR / fname
            exporter = FBAExporter(template_path=tpl_path, tracking_map=tracking_map)
            exporter.export(woven, str(out))

            DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
            shutil.copy2(out, DOWNLOADS_DIR / fname)

            self._json({
                "success": True,
                "download_url": f"/downloads/{fname}",
                "summary": {
                    "shipment_id": woven["shipment_id"],
                    "total_boxes": woven["total_boxes"],
                    "total_weight": woven["total_weight"],
                    "total_cbm": woven["total_cbm"],
                    "row_count": len(woven["rows"]),
                },
                "rows": woven["rows"],
            })
        except Exception as e:
            self._json({"success": False, "error": str(e)}, 400)

    def _handle_weave_batch(self):
        """批量上传多个 FBA 货件，一次性编织并返回下载链接 + zip base64"""
        import subprocess as _sp, sys as _sys
        try:
            body = json.loads(self._read())
            files = body.get("files", [])
            hs_code = body.get("hs_code", None)
            tpl_b64 = body.get("template", None)
            track_b64 = body.get("tracking", None)
            if not files:
                self._json({"success": False, "error": "未提供文件"}, 400)
                return
            tpl_path = None
            if tpl_b64:
                tpl_path = self._save(tpl_b64, "tpl")
            tracking_map = None
            if track_b64:
                tracking_map = self._parse_tracking(self._save(track_b64, "track"))
            parser = FBAParser()
            exporter = FBAExporter(template_path=tpl_path, tracking_map=tracking_map)
            DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            results = []
            for f in files:
                b64 = f.get("data", "")
                fname_orig = f.get("name", "未知")
                if not b64:
                    results.append({"filename": fname_orig, "success": False, "error": "文件数据为空"})
                    continue
                try:
                    path = self._save(b64, "fba")
                    data = parser.parse(path)
                    meta = data.get("meta", {})
                    sid = meta.get("shipment_id", "UNKNOWN")
                    woven = weave_fba(data, hs_code_override=hs_code)
                    out_fname = f"{sid}-装箱单_{uuid.uuid4().hex[:6]}.xlsx"
                    out = UPLOAD_DIR / out_fname
                    exporter.export(woven, str(out))
                    shutil.copy2(out, DOWNLOADS_DIR / out_fname)
                    results.append({
                        "filename": fname_orig, "success": True,
                        "shipment_id": sid,
                        "download_url": f"/downloads/{out_fname}",
                        "summary": {
                            "shipment_id": woven["shipment_id"],
                            "total_boxes": woven["total_boxes"],
                            "total_weight": woven["total_weight"],
                            "total_cbm": woven["total_cbm"],
                            "row_count": len(woven["rows"]),
                        },
                    })
                except Exception as e:
                    results.append({"filename": fname_orig, "success": False, "error": str(e)})
            # 生成 zip base64（内嵌在响应中，前端直接解码下载）
            zip_b64 = ""
            ok = [r for r in results if r.get("success")]
            if ok:
                try:
                    bid = uuid.uuid4().hex[:6]
                    td = UPLOAD_DIR / f"batch_{bid}"
                    td.mkdir(parents=True, exist_ok=True)
                    for r in ok:
                        src = DOWNLOADS_DIR / Path(r["download_url"]).name
                        if src.exists():
                            shutil.copy2(src, td / src.name)
                    zd = str(UPLOAD_DIR / f"batch_{bid}.zip")
                    try:
                        _sp.run(["zip", "-j", zd] + sorted([f.name for f in td.iterdir()]),
                                 cwd=str(td), check=True, capture_output=True, timeout=30)
                    except (_sp.CalledProcessError, FileNotFoundError):
                        with zipfile.ZipFile(zd, "w") as zf:
                            for fp in sorted(td.iterdir()):
                                zf.write(str(fp), fp.name)
                    zip_b64 = base64.b64encode(Path(zd).read_bytes()).decode()
                    shutil.rmtree(td, ignore_errors=True)
                    Path(zd).unlink(missing_ok=True)
                except Exception:
                    pass
            self._json({"success": True, "results": results, "zip_b64": zip_b64})
        except Exception as e:
            self._json({"success": False, "error": str(e)}, 400)

    def _handle_merge_plan(self):
        """上传调整明细 + 发货计划，按四步规则调整后输出"""
        try:
            body = json.loads(self._read())
            b64_adj = body.get("file", "")
            b64_plan = body.get("plan_file", "")
            if not b64_adj or not b64_plan:
                self._json({"success": False, "error": "请上传调整明细和发货计划"}, 400)
                return

            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

            # 保存调整明细
            adj_path = UPLOAD_DIR / f"adj_{uuid.uuid4().hex[:6]}.xlsx"
            adj_path.write_bytes(base64.b64decode(b64_adj))
            # 保存发货计划
            plan_path = UPLOAD_DIR / f"plan_{uuid.uuid4().hex[:6]}.xlsx"
            plan_path.write_bytes(base64.b64decode(b64_plan))

            # ====== 解析调整明细（按表头名称定位列） ======
            wb_adj = openpyxl.load_workbook(str(adj_path), data_only=True)
            ws_adj = wb_adj.active
            col_idx = {}
            for c in range(1, ws_adj.max_column + 1):
                h = str(ws_adj.cell(row=1, column=c).value or "")
                if h.strip() == "SKU": col_idx["SKU"] = c
                elif "调整单号" in h: col_idx["调整单号"] = c
                elif "识别码" in h and "新" not in h: col_idx["识别码"] = c
                elif "调整FNSKU" in h: col_idx["调整FNSKU"] = c
                elif "调整量" in h: col_idx["调整量"] = c
                elif "调整店铺" in h: col_idx["调整店铺"] = c
                elif "店铺" in h and "调整" not in h: col_idx["店铺"] = c
            if "SKU" not in col_idx or "识别码" not in col_idx:
                wb_adj.close(); wb_plan.close()
                self._json({"success": False, "error": "未找到SKU或识别码列"}, 400)
                return
            adjustments = []
            for r in range(2, ws_adj.max_row + 1):
                code = ws_adj.cell(row=r, column=col_idx["识别码"]).value
                if code is None: continue
                qty = ws_adj.cell(row=r, column=col_idx.get("调整量",0)).value
                adjustments.append({
                    "调整单号": str(ws_adj.cell(row=r, column=col_idx.get("调整单号",0)).value or "").strip(),
                    "识别码": str(code).strip(),
                    "店铺": str(ws_adj.cell(row=r, column=col_idx.get("店铺",0)).value or "").strip(),
                    "SKU": str(ws_adj.cell(row=r, column=col_idx["SKU"]).value or "").strip(),
                    "调整FNSKU": str(ws_adj.cell(row=r, column=col_idx.get("调整FNSKU",0)).value or "").strip(),
                    "调整量": int(qty) if qty is not None else 0,
                    "调整店铺": str(ws_adj.cell(row=r, column=col_idx.get("调整店铺",0)).value or "").strip(),
                })
            wb_adj.close()

            if not adjustments:
                self._json({"success": False, "error": "调整明细无有效数据"}, 400)
                return

            # ====== 打开发货计划（用户上传的文件） ======
            wb_plan = openpyxl.load_workbook(str(plan_path))
            ws_plan = wb_plan.active

            # 解析计划表
            plan_rows = []
            for r in range(2, ws_plan.max_row + 1):
                plan_rows.append({
                    "r": r,
                    "识别码": str(ws_plan.cell(row=r, column=3).value or "").strip(),
                    "FNSKU": str(ws_plan.cell(row=r, column=6).value or "").strip(),
                    "店铺": str(ws_plan.cell(row=r, column=15).value or "").strip(),
                    "计划发货量": int(ws_plan.cell(row=r, column=9).value) if ws_plan.cell(row=r, column=9).value else 0,
                    "箱数": int(ws_plan.cell(row=r, column=8).value) if ws_plan.cell(row=r, column=8).value else 0,
                    "is_new": (ws_plan.cell(row=r, column=1).value is None),
                })

            # ====== 规则13: 筛选同店铺 + 规则14: 识别码+FNSKU匹配 ======
            match_map = {}
            for adj in adjustments:
                for pr in plan_rows:
                    if adj["调整店铺"] == pr["店铺"] and adj["识别码"] == pr["识别码"] and adj["调整FNSKU"] == pr["FNSKU"]:
                        key = (adj["识别码"], adj["调整FNSKU"])
                        if key not in match_map:
                            match_map[key] = {"plan_row": pr, "adjustments": [], "total_adj_qty": 0}
                        match_map[key]["adjustments"].append(adj)
                        match_map[key]["total_adj_qty"] += adj["调整量"]
                        break

            # ====== 规则15: 数量判断 + 规则16: 写入和清理 ======
            yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
            deleted = set()

            for key, data in match_map.items():
                pr = data["plan_row"]
                adj_list = data["adjustments"]
                r = pr["r"]
                adj_nos = ",".join(sorted(set(a["调整单号"] for a in adj_list)))

                # 填入调整单号（列17=调拨单号/调拨量）
                ws_plan.cell(row=r, column=17).value = adj_nos
                # 更新国家列为调整店铺
                ws_plan.cell(row=r, column=4).value = adj_list[0]["调整店铺"]
                # 补库存箱数（覆盖公式或空值）
                v = ws_plan.cell(row=r, column=11).value
                if v is None or (isinstance(v, str) and v.startswith("=")):
                    ws_plan.cell(row=r, column=11).value = pr["箱数"]

            # 删除已处理的黄色新增行
            for pr in plan_rows:
                if pr["is_new"]:
                    key = (pr["识别码"], pr["FNSKU"])
                    if key in match_map:
                        for c in range(1, ws_plan.max_column + 1):
                            ws_plan.cell(row=pr["r"], column=c).value = None
                        deleted.add(pr["r"])

            # 补全所有行的库存箱数（覆盖公式或空值）
            for pr in plan_rows:
                if pr["r"] not in deleted and pr["箱数"] > 0:
                    v = ws_plan.cell(row=pr["r"], column=11).value
                    if v is None or (isinstance(v, str) and v.startswith("=")):
                        ws_plan.cell(row=pr["r"], column=11).value = pr["箱数"]

            # 保存输出
            out_name = f"发货计划_{uuid.uuid4().hex[:6]}.xlsx"
            out_path = UPLOAD_DIR / out_name
            wb_plan.save(str(out_path))
            wb_plan.close()

            DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
            shutil.copy2(out_path, DOWNLOADS_DIR / out_name)

            self._json({
                "success": True,
                "download_url": f"/downloads/{out_name}",
                "matched": len(match_map),
                "deleted": len(deleted),
            })
        except Exception as e:
            self._json({"success": False, "error": str(e)}, 400)

    def _handle_download_zip(self):
        """将所有已生成的装箱单打包为 zip 下载"""
        import subprocess, sys, traceback
        try:
            DOWNLOADS_DIR.mkdir(parents=True, exist_ok=True)
            xlsx_files = sorted(DOWNLOADS_DIR.glob("*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
            if not xlsx_files:
                self.send_error(404)
                return

            UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
            batch_id = uuid.uuid4().hex[:6]
            tmp_dir = UPLOAD_DIR / f"batch_{batch_id}"
            tmp_dir.mkdir(parents=True, exist_ok=True)
            for fp in xlsx_files:
                shutil.copy2(fp, tmp_dir / fp.name)

            zip_dest = str(UPLOAD_DIR / f"batch_{batch_id}.zip")
            try:
                subprocess.run(
                    ["zip", "-j", zip_dest] + sorted([f.name for f in tmp_dir.iterdir()]),
                    cwd=str(tmp_dir), check=True, capture_output=True, timeout=30
                )
            except (subprocess.CalledProcessError, FileNotFoundError):
                with zipfile.ZipFile(zip_dest, "w") as zf:
                    for fp in sorted(tmp_dir.iterdir()):
                        zf.write(str(fp), fp.name)

            data = Path(zip_dest).read_bytes()
            shutil.rmtree(tmp_dir, ignore_errors=True)
            Path(zip_dest).unlink(missing_ok=True)

            self.send_response(200)
            self.send_header("Content-Type", "application/zip")
            self.send_header("Content-Disposition", "attachment; filename=装箱单_批量.zip")
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            traceback.print_exc(file=sys.stderr)
            self.send_error(500)


def run_server(port: int = 8888):
    import socketserver
    class ReusableServer(HTTPServer):
        allow_reuse_address = True
    server = ReusableServer(("0.0.0.0", port), BaoHandler)
    print(f"🚀 bao Web 面板 — http://127.0.0.1:{port}")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n🛑 服务已停止")
        server.server_close()
