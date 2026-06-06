"""FBA 装箱单 Excel 导出器 — 模板颜色分灰/黄 + DISPIMG 支持"""
import shutil
import tempfile
from pathlib import Path
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Alignment, Font

_DEFAULT_TEMPLATE = Path(__file__).parent.parent.parent / "templates" / "德速-模板.xlsx"

PATTERNS = [
    ("shipment_id",  ["shipment id", "shipmentid"]),
    ("reference_id", ["reference id", "referenceid"]),
    ("reference_no", ["参考号"]),
    ("box_range",    ["箱号段", "箱号"]),
    ("total_boxes",  ["总件数"]),
    ("warehouse",    ["目的仓库代码", "仓库代码"]),
    ("channel",      ["渠道"]),
    ("sku",          ["sku"]),
    ("en_name",      ["英文品名"]),
    ("cn_name",      ["中文品名"]),
    ("brand",        ["brand", "品牌"]),
    ("material",     ["材质"]),
    ("usage",        ["用途"]),
    ("hs_code",      ["海关编码", "hs", "hscode", "进口海关编码"]),
    ("asin_url",     ["asin", "销售链接"]),
    ("battery",      ["带电"]),
    ("model",        ["型号"]),
    ("total_qty",    ["总数量", "每箱数量"]),
    ("unit",         ["单位"]),
    ("per_set",      ["每套"]),
    ("cost",         ["采购单价", "采购", "投保单价"]),
    ("declare_price",["申报单价"]),
    ("image",        ["图片"]),
    ("box_weight",   ["单箱重量", "重量", "单箱重"]),
    ("length",       ["长", "单箱长"]),
    ("width",        ["宽", "单箱宽"]),
    ("height",       ["高", "单箱高"]),
]

YELLOW_FILL = openpyxl.styles.PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
GRAY_FILL = openpyxl.styles.PatternFill(start_color="FFD8D8D8", end_color="FFD8D8D8", fill_type="solid")


import logging

class FBAExporter:
    def __init__(self, template_path: str = None, tracking_map: dict = None, shipping_map: dict = None):
        self._template_path = Path(template_path) if template_path else _DEFAULT_TEMPLATE
        raw = tracking_map or {}
        self._tracking_map = {}
        for k, v in raw.items():
            if isinstance(v, dict):
                self._tracking_map[k] = v
            else:
                self._tracking_map[k] = {"tracking_code": str(v)}
        if shipping_map:
            for k, v in shipping_map.items():
                if k in self._tracking_map:
                    if isinstance(v, dict):
                        for fk in ("warehouse", "channel", "total_boxes"):
                            if fk in v and fk not in self._tracking_map[k]:
                                self._tracking_map[k][fk] = v[fk]
                else:
                    if isinstance(v, dict):
                        self._tracking_map[k] = v
                    else:
                        self._tracking_map[k] = {"tracking_code": str(v)}
        if not self._template_path.exists():
            raise FileNotFoundError(f"模板不存在: {self._template_path}")

    def export(self, woven: dict, output_path: str) -> str:
        shutil.copy2(self._template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws_name = "下单模板" if "下单模板" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[ws_name]
        rows = woven.get("rows", [])

        header_row, col_map, meta_cells = self._detect_template(ws)
        data_start = header_row + 1

        def col(field: str) -> int:
            return col_map.get(field, 0)

        # ── 灰/黄分类 ──────────────────────────────────
        grey_fields = set()
        for c in range(1, min(30, ws.max_column or 0) + 1):
            cell = ws.cell(row=header_row, column=c)
            try:
                fg = cell.fill.fgColor
                is_grey = (fg.type == 'theme' and fg.theme is not None)
            except Exception:
                is_grey = False
            if not is_grey:
                continue
            val = str(cell.value or "").strip().lower()
            for field, kws in PATTERNS:
                for kw in kws:
                    if kw in val and field not in grey_fields:
                        grey_fields.add(field)
                        break

        # ── 提取 cellimages (DISPIMG) ──────────────────
        dispimg_info = self._extract_cell_images(str(output_path))

        # ── 九方：提取模板嵌入图（drawing1.xml）─────────────
        _is_jiufang = "九方" in str(self._template_path)
        _img_by_row = {}  # {template_row: (bytes, w, h)}
        if _is_jiufang and not dispimg_info:
            for img in ws._images:
                try:
                    row_idx = None
                    if hasattr(img.anchor, '_from'):
                        row_idx = img.anchor._from.row + 1
                    if row_idx is not None:
                        _img_by_row[row_idx] = (img._data(), img.width or 54, img.height or 66)
                except Exception:
                    continue

        # ── 提取 openpyxl 嵌入图（cellimages 不存在时）───
        t_img_bytes, t_w, t_h = None, 54, 66
        if not dispimg_info and not _img_by_row:
            for img in ws._images:
                try:
                    t_img_bytes = img._data()
                    t_w = img.width or 54
                    t_h = img.height or 66
                    break
                except Exception:
                    continue
        if not t_img_bytes and not dispimg_info and not _img_by_row and _DEFAULT_TEMPLATE.exists():
            try:
                dwb = openpyxl.load_workbook(_DEFAULT_TEMPLATE)
                dsn = "下单模板" if "下单模板" in dwb.sheetnames else dwb.sheetnames[0]
                for img in dwb[dsn]._images:
                    try:
                        t_img_bytes = img._data()
                        t_w = img.width or 54
                        t_h = img.height or 66
                        break
                    except Exception:
                        continue
                dwb.close()
            except Exception:
                pass

        # ── HS 品类映射 + DISPIMG 关联（仅需从模板复制的字段）───
        # 从模板复制的字段：海关编码/是否带电/型号/单位/每套个数
        _GREY_COPY_FIELDS = {"battery", "model", "unit", "per_set"}
        # 从模板第一个数据行提取默认值（无需按 HS 匹配）
        _grey_defaults = {}
        for fld in _GREY_COPY_FIELDS:
            c = col(fld)
            if c > 0:
                v = ws.cell(row=data_start, column=c).value
                if v is not None:
                    _grey_defaults[fld] = str(v)
        # 海关编码也从模板第一行复制
        if col("hs_code") > 0:
            v = ws.cell(row=data_start, column=col("hs_code")).value
            if v is not None:
                _grey_defaults["hs_code"] = str(v).strip()
        # DISPIMG 取模板第一个图片（不按 HS 匹配）
        _first_dispimg = None
        for drow, (dname, _ib, _w, _h) in dispimg_info.items():
            if drow == data_start:
                _first_dispimg = dname
                break
        if _first_dispimg is None and dispimg_info:
            _first_dispimg = list(dispimg_info.values())[0][0]

        template_categories = {}
        if col("hs_code") > 0:
            for sr in range(data_start, min(data_start + 10, ws.max_row + 1)):
                hs = ws.cell(row=sr, column=col("hs_code")).value
                if not hs:
                    continue
                hs_key = str(hs).strip()
                if hs_key in template_categories:
                    continue
                cat_data = {}
                for fld in grey_fields:
                    c = col(fld)
                    if c > 0:
                        v = ws.cell(row=sr, column=c).value
                        if v is not None:
                            cat_data[fld] = str(v)
                for drow, (dname, _ib, _w, _h) in dispimg_info.items():
                    if drow == sr:
                        cat_data["_dispimg"] = dname
                        break
                if cat_data:
                    template_categories[hs_key] = cat_data

        # ── 九方模板：按中文品名匹配品类（蜡烛/香氛香薰）─────
        _JIUFANG_FIELDS = {"hs_code", "en_name", "cn_name", "brand", "material", "usage",
                           "battery", "model", "unit", "per_set"}
        _JIUFANG_OVERRIDE_FIELDS = _JIUFANG_FIELDS - _GREY_COPY_FIELDS
        _jiufang_xiangxun = None
        _jiufang_lazhu = None
        _jiufang_fallback = None  # 极端兜底：模板第一个有数据的行
        _jiufang_row_img = []  # [(data_row_index, tpl_row)]

        if _is_jiufang and col("cn_name") > 0:
            for sr in range(data_start, min(data_start + 10, ws.max_row + 1)):
                cn = str(ws.cell(row=sr, column=col("cn_name")).value or "").strip()
                if not cn:
                    continue
                cat_data = {}
                for fld in _JIUFANG_FIELDS:
                    c = col(fld)
                    if c > 0:
                        v = ws.cell(row=sr, column=c).value
                        if v is not None:
                            cat_data[fld] = str(v)
                for drow, (dname, _ib, _w, _h) in dispimg_info.items():
                    if drow == sr:
                        cat_data["_dispimg"] = dname
                        break
                # 嵌入图（drawing1.xml）
                if sr in _img_by_row:
                    cat_data["_embedded_img"] = _img_by_row[sr] + (sr,)
                if "香薰" in cn or "香氛" in cn:
                    _jiufang_xiangxun = cat_data
                elif "蜡烛" in cn:
                    _jiufang_lazhu = cat_data
                # 记录第一个有中文品名的行作为极端兜底
                if _jiufang_fallback is None:
                    _jiufang_fallback = cat_data
            if _jiufang_lazhu is None:
                for sr in range(data_start, min(data_start + 10, ws.max_row + 1)):
                    cn = str(ws.cell(row=sr, column=col("cn_name")).value or "").strip()
                    # 跳过空行和香薰行
                    if not cn or "香薰" in cn or "香氛" in cn:
                        continue
                    cat_data = {}
                    for fld in _JIUFANG_FIELDS:
                        c = col(fld)
                        if c > 0:
                            v = ws.cell(row=sr, column=c).value
                            if v is not None:
                                cat_data[fld] = str(v)
                    for drow, (dname, _ib, _w, _h) in dispimg_info.items():
                        if drow == sr:
                            cat_data["_dispimg"] = dname
                            break
                    if sr in _img_by_row:
                        cat_data["_embedded_img"] = _img_by_row[sr] + (sr,)
                    _jiufang_lazhu = cat_data
                    break

        # ── 保存模板边框样式（九方用）──────────────────
        from copy import copy
        from openpyxl.styles import Border as _Border, Side as _Side
        _ref_border = None
        if _is_jiufang and data_start <= ws.max_row:
            ref_cell = ws.cell(row=data_start, column=1)
            if not isinstance(ref_cell, MergedCell):
                rb = ref_cell.border
                _ref_border = _Border(
                    left=copy(rb.left), right=copy(rb.right),
                    top=copy(rb.top), bottom=copy(rb.bottom))

        # ── 清除数据行 ─────────────────────────────────
        max_c = col_map.get("_max_col", 24)
        for r in range(data_start, data_start + 200):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                if not isinstance(cell, MergedCell):
                    cell.value = None
        ws._images.clear()

        r2 = {"total_boxes": woven.get("total_boxes", 0),
              "total_cbm": woven.get("total_cbm", 0),
              "total_weight": woven.get("total_weight", 0)}
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        bf = Font(name="微软雅黑", size=10)

        for fld, val in r2.items():
            c = col(fld)
            if c > 0:
                ws.cell(row=2, column=c, value=val)

        # 从 tracking_map 填充模板元数据单元格（参考号/仓库代码/渠道/总件数）
        if meta_cells and self._tracking_map:
            sid = woven.get("shipment_id", "")
            info = self._tracking_map.get(sid, {})
            for field, (mr, mc) in meta_cells.items():
                val = None
                if field == "reference_no":
                    val = sid
                elif field == "total_boxes":
                    val = info.get("total_boxes")
                elif field == "warehouse":
                    val = info.get("warehouse")
                elif field == "channel":
                    val = info.get("channel")
                elif field == "shipment_id":
                    val = sid
                if val is not None:
                    self._set_col(ws, mr, mc, str(val), center, bf, YELLOW_FILL)

        # ── 九方：币种写入 Row2 S列 ──
        if _is_jiufang:
            _cc = woven.get("country_code", "")
            _currency_map = {"US": "USD", "CA": "USD", "AU": "USD", "UK": "GBP", "DE": "EUR"}
            _currency = _currency_map.get(_cc, "")
            if _currency and col("declare_price") > 0:
                # 找申报单价列（实际币种写在附近，这里写备注或直接改Row2某列）
                # 模板用 Row2 S列 = 申报币种*, 覆盖为实际币种
                try:
                    ccy_cell = ws.cell(row=2, column=col("declare_price"))
                    if not isinstance(ccy_cell, MergedCell):
                        ccy_cell.value = _currency
                except: pass
        _temp_files = []

        for i, rd in enumerate(rows):
            r = data_start + i
            # 每行独立：已被九方品类数据覆盖的字段（后续黄色/灰色写入需跳过）
            _jiufang_written = set()
            # 九方模板保持原有行高，不强制设置
            if not _is_jiufang:
                ws.row_dimensions[r].height = 40

            row_hs = str(rd.get("进口海关编码") or rd.get("海关编码", "")).strip()
            row_cn = str(rd.get("中文品名", "")).strip()
            tmpl_cat = template_categories.get(row_hs) if row_hs else None

            # ── 九方：按中文品名选择品类 ──────────────────
            cat = None
            if _is_jiufang:
                _ship_name = woven.get("shipment_name", "")
                _product_name = str(rd.get("标题", ""))
                if ("香薰" in row_cn or "香氛" in row_cn or "挥发液" in _ship_name or "挥发液" in _product_name) and _jiufang_xiangxun:
                    cat = _jiufang_xiangxun
                else:
                    cat = _jiufang_lazhu or _jiufang_fallback
                # 九方数据覆盖（品类字段写入）
                _override_source = cat if cat else tmpl_cat
                if _override_source:
                    for fld in _JIUFANG_OVERRIDE_FIELDS:
                        if fld not in _override_source:
                            continue
                        c = col(fld)
                        if c > 0:
                            self._set_col(ws, r, c, _override_source[fld], center, bf, None)
                            _jiufang_written.add(fld)
                    # 灰色字段也从品类数据取值
                    # 九方规则：Brand 固定为 M&SENSE
                    if col("brand") > 0:
                        self._set_col(ws, r, col("brand"), "M&SENSE", center, bf, None)
                        _jiufang_written.add("brand")
                    for fld in _GREY_COPY_FIELDS:
                        if fld not in _override_source:
                            continue
                        c = col(fld)
                        if c > 0:
                            self._set_col(ws, r, c, _override_source[fld], center, bf, None)
                            _jiufang_written.add(fld)

            # 从模板复制灰色字段（所有检测到的灰色列）
            for fld in grey_fields:
                c = col(fld)
                if c <= 0:
                    continue
                # 九方模式下已由品类数据写入，跳过
                if _is_jiufang and fld in _jiufang_written:
                    continue
                val = None
                if cat and fld in cat:
                    val = cat[fld]
                elif tmpl_cat and fld in tmpl_cat:
                    val = tmpl_cat[fld]
                if val is None:
                    val = _grey_defaults.get(fld)
                if val is not None:
                    self._set_col(ws, r, c, str(val), center, bf, GRAY_FILL)

            # 货件追踪码 lookup
            if col("reference_id") > 0 and self._tracking_map:
                sid = woven.get("shipment_id", "")
                info = self._tracking_map.get(sid, {})
                track_code = info.get("tracking_code", "")
                if track_code:
                    self._set_col(ws, r, col("reference_id"), track_code, center, bf, None)
                elif r == data_start:
                    import logging
                    logging.getLogger("bao").warning(
                        f"追踪码未匹配: sid={sid!r}, map_keys={list(self._tracking_map.keys())[:5]}"
                    )

            yw = {
                "shipment_id": woven.get("shipment_id", ""),
                "box_range":   rd.get("箱号段", ""),
                "total_boxes": int(rd.get("总件数", 0)),
                "sku":         rd.get("SKU", ""),
                "total_qty":   int(rd.get("总数量", 0)),
                "box_weight":  rd.get("单箱重量", 0),
                "length":      rd.get("长", 0),
                "width":       rd.get("宽", 0),
                "height":      rd.get("高", 0),
                "en_name":     rd.get("英文品名", ""),
                "cn_name":     rd.get("中文品名", ""),
                "material":    rd.get("材质", ""),
                "usage":       rd.get("用途", ""),
                "brand":       rd.get("品牌", ""),
                "hs_code":     rd.get("海关编码", ""),
            }
            for fld, val in yw.items():
                # 九方模式下已被品类数据覆盖的字段，跳过不再写入
                if _is_jiufang and fld in _jiufang_written:
                    continue
                if fld in grey_fields or val is None:
                    continue
                c = col(fld)
                if c > 0:
                    _fill = YELLOW_FILL if not _is_jiufang else None
                    self._set_col(ws, r, c, val, center, bf, _fill)

            if col("asin_url") > 0 and "asin_url" not in grey_fields:
                self._set_col(ws, r, col("asin_url"),
                              str(rd.get("ASIN", "")), center, bf, GRAY_FILL)

            # ── 九方：采购单价=申报单价*7 公式 ──
            if _is_jiufang and col("cost") > 0 and col("declare_price") > 0:
                declare_col_letter = openpyxl.utils.get_column_letter(col("declare_price"))
                declare_cell = f"{declare_col_letter}{r}"
                ws.cell(row=r, column=col("cost")).value = f"={declare_cell}*7"
                _jiufang_written.add("cost")
            # ── 九方：ASIN 留空 ──
            if _is_jiufang and col("asin_url") > 0:
                ws.cell(row=r, column=col("asin_url")).value = ""
                _jiufang_written.add("asin_url")
            # ── 图片 ──────────────────────────────────
            if col("image") > 0:
                if _is_jiufang:
                    # 九方：记录品类行映射，由 _inject_drawing_images 后处理注入
                    _img_source = cat if cat else tmpl_cat
                    if _img_source and "_embedded_img" in _img_source:
                        _tpl_row = _img_source["_embedded_img"][3]
                        _jiufang_row_img.append((i, _tpl_row))
                else:
                    # 德速/通用：DISPIMG 公式 or TwoCellAnchor 嵌入
                    disp_name = None
                    if tmpl_cat and "_dispimg" in tmpl_cat:
                        disp_name = tmpl_cat["_dispimg"]
                    elif _first_dispimg:
                        disp_name = _first_dispimg
                    if disp_name:
                        ws.cell(row=r, column=col("image")).value = f'=_xlfn.DISPIMG("{disp_name}",1)'
                    elif t_img_bytes:
                        tf = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
                        tf.write(t_img_bytes)
                        tf.close()
                        _temp_files.append(tf.name)
                        img = XLImage(tf.name)
                        img.width, img.height = max(t_w, 54), max(t_h, 66)
                        img.anchor = TwoCellAnchor(
                            _from=AnchorMarker(col=col("image") - 1, row=r - 1,
                                               colOff=0, rowOff=0),
                            to=AnchorMarker(col=col("image"), row=r, colOff=0, rowOff=0),
                            editAs="oneCell")
                        ws.add_image(img)

        hf = Font(name="微软雅黑", bold=True, size=11)
        for c in range(1, max_c + 1):
            cell = ws.cell(row=2, column=c)
            if cell.value is not None:
                cell.font = hf

        # ── 九方：统一数据行边框与字号（与模板一致）─────
        if _is_jiufang:
            from openpyxl.styles import Border as _Bdr, Side as _Sd
            _thin_side = _Sd(style='thin', color='FF000000')
            _data_border = _Bdr(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)
            _data_font = Font(name='宋体', size=11)
            # 统一行高取模板数据首行（避免新增行用默认行高导致图片压缩）
            _std_row_h = ws.row_dimensions[data_start].height
            if _std_row_h is None:
                _std_row_h = 50
            for r in range(data_start, data_start + len(rows)):
                ws.row_dimensions[r].height = _std_row_h
                for c in range(1, max_c + 1):
                    cell = ws.cell(row=r, column=c)
                    if not isinstance(cell, MergedCell):
                        cell.border = _data_border
                        cell.font = _data_font
                        cell.alignment = center

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        wb.close()

        # ── 九方：后处理注入嵌入图 + 统一边框 ──────────
        if _is_jiufang:
            if _jiufang_row_img:
                self._inject_drawing_images(output_path, data_start, _jiufang_row_img, col("image"))


        for tf in _temp_files:
            try:
                Path(tf).unlink()
            except Exception:
                pass

        if dispimg_info:
            self._inject_cellimages(output_path)

        return output_path

    def _inject_drawing_images(self, output_path: str, data_start: int, row_img_map: list, image_col: int):
        """九方专用：按 row_img_map [(data_idx, tpl_row), ...] 将模板嵌入图注入输出 xlsx"""
        if image_col <= 0 or not row_img_map:
            return
        import zipfile as _zf, os as _os, shutil as _sh, re as _re, xml.etree.ElementTree as _ET
        tmp_path = output_path + ".tmp"
        try:
            ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
            ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
            ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
            with _zf.ZipFile(output_path, 'r') as zin:
                tpl_draw = None
                tpl_rels = None
                tpl_media = {}
                with _zf.ZipFile(str(self._template_path), 'r') as tpl:
                    if 'xl/drawings/drawing1.xml' in tpl.namelist():
                        tpl_draw = tpl.read('xl/drawings/drawing1.xml')
                    if 'xl/drawings/_rels/drawing1.xml.rels' in tpl.namelist():
                        tpl_rels = tpl.read('xl/drawings/_rels/drawing1.xml.rels')
                    for name in tpl.namelist():
                        if name.startswith('xl/media/'):
                            tpl_media[name] = tpl.read(name)
                if not tpl_draw:
                    return
                draw_root = _ET.fromstring(tpl_draw)
                xdr_tag = f'{{{ns_xdr}}}'
                row_pics = {}
                for anchor in draw_root.findall(f'{xdr_tag}twoCellAnchor'):
                    from_el = anchor.find(f'{xdr_tag}from')
                    if from_el is None: continue
                    row_el = from_el.find(f'{xdr_tag}row')
                    if row_el is None: continue
                    tpl_row = int(row_el.text) + 1
                    row_pics.setdefault(tpl_row, []).append(anchor)
                new_root = _ET.Element(f'{xdr_tag}wsDr', {'xmlns:xdr': ns_xdr, 'xmlns:a': ns_a})
                new_media = {}
                new_rels = _ET.Element('Relationships', {'xmlns': ns_r})
                rid_counter = 1
                for data_idx, tpl_row in row_img_map:
                    if tpl_row not in row_pics:
                        continue
                    r = data_start + data_idx
                    for pic in row_pics[tpl_row]:
                        new_pic = _ET.fromstring(_ET.tostring(pic))
                        for el in new_pic.iter(f'{xdr_tag}row'):
                            el.text = str(r - 1)
                        for el in new_pic.iter(f'{xdr_tag}col'):
                            el.text = str(image_col - 1)
                        for blip in new_pic.iter(f'{{{ns_a}}}blip'):
                            old_rid = blip.get(f'{{{ns_r}}}embed')
                            if old_rid and tpl_rels:
                                rels_root = _ET.fromstring(tpl_rels)
                                for rel in rels_root:
                                    if rel.get('Id') == old_rid:
                                        target = rel.get('Target')
                                        new_rid = f'rId{rid_counter}'
                                        blip.set(f'{{{ns_r}}}embed', new_rid)
                                        _ET.SubElement(new_rels, 'Relationship', {
                                            'Id': new_rid,
                                            'Type': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/image',
                                            'Target': target,
                                        })
                                        # target 是 '../media/xxx.png'，规范化到 'xl/media/xxx.png'
                                        media_key = 'xl/media/' + target.rsplit('/', 1)[-1]
                                        if media_key in tpl_media:
                                            new_media[media_key] = tpl_media[media_key]
                                        rid_counter += 1
                                        break
                        new_root.append(new_pic)
                new_draw = _ET.tostring(new_root, encoding='unicode')
                new_rels_xml = _ET.tostring(new_rels, encoding='unicode')
                # ── 构建 sheet1.xml.rels（drawing 关联到"下单模板"sheet）──
                sheet1_rels = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>'

                # ── 修改 sheet1.xml 注入 <drawing r:id="rId1"/> ──
                sheet1_xml = zin.read('xl/worksheets/sheet1.xml').decode()
                if '<drawing' not in sheet1_xml:
                    sheet1_xml = sheet1_xml.replace('</worksheet>', '<drawing r:id="rId1" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"/></worksheet>')

                # ── 清理其他 sheet 的 drawing 关联 ──
                sheet2_xml = None
                sheet2_rels = None
                if 'xl/worksheets/sheet2.xml' in {i.filename for i in zin.infolist()}:
                    sheet2_xml = zin.read('xl/worksheets/sheet2.xml').decode()
                    if '<drawing' in sheet2_xml:
                        sheet2_xml = _re.sub(r'<drawing[^>]*/>', '', sheet2_xml)
                if 'xl/worksheets/_rels/sheet2.xml.rels' in {i.filename for i in zin.infolist()}:
                    sheet2_rels = zin.read('xl/worksheets/_rels/sheet2.xml.rels').decode()
                    if 'drawing' in sheet2_rels:
                        sheet2_rels = _re.sub(r'<Relationship[^>]*drawing[^>]*/>', '', sheet2_rels)

                # ── 清理 workbook.xml.rels 中的 drawing 引用 ──
                wb_rels = zin.read('xl/_rels/workbook.xml.rels').decode()
                wb_rels = _re.sub(r'<Relationship[^>]*drawing[^>]*/>', '', wb_rels)

                with _zf.ZipFile(tmp_path, 'w', _zf.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        if 'drawing1.xml' in item.filename:
                            continue
                        if item.filename == 'xl/_rels/workbook.xml.rels':
                            zout.writestr(item, wb_rels)
                        elif item.filename == 'xl/worksheets/sheet1.xml':
                            zout.writestr(item, sheet1_xml)
                        elif item.filename == 'xl/worksheets/sheet2.xml' and sheet2_xml is not None:
                            zout.writestr(item, sheet2_xml)
                        elif item.filename == 'xl/worksheets/_rels/sheet2.xml.rels' and sheet2_rels is not None:
                            zout.writestr(item, sheet2_rels)
                        elif item.filename == '[Content_Types].xml':
                            ct = zin.read(item.filename).decode()
                            if 'drawing' not in ct:
                                ct = ct.replace('</Types>', '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
                            zout.writestr(item, ct)
                        else:
                            zout.writestr(item, zin.read(item.filename))
                    # 写入 sheet1.xml.rels（总是新建或覆盖）
                    zout.writestr('xl/worksheets/_rels/sheet1.xml.rels', sheet1_rels)
                    zout.writestr('xl/drawings/drawing1.xml', new_draw)
                    zout.writestr('xl/drawings/_rels/drawing1.xml.rels', new_rels_xml)
                    for mp, md in new_media.items():
                        if mp not in {i.filename for i in zin.infolist()}:
                            zout.writestr(mp, md)
            _sh.move(tmp_path, output_path)
        except Exception as _ie:
            import traceback as _tb
            _tb.print_exc()
            if _os.path.exists(tmp_path):
                _os.remove(tmp_path)

    def _detect_template(self, ws) -> tuple:
        max_col = min(30, ws.max_column or 0)
        best_row, best_score = 0, 0
        for r in range(1, min(31, ws.max_row + 1)):
            score = 0
            for c in range(1, max_col + 1):
                val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                for _, kws in PATTERNS:
                    for kw in kws:
                        if kw in val:
                            score += 1
                            break
            if score > best_score:
                best_score, best_row = score, r
        if best_row == 0:
            best_row = 3
        col_map = {"_max_col": max_col}
        for c in range(1, max_col + 1):
            val = str(ws.cell(row=best_row, column=c).value or "").strip().lower()
            if not val:
                continue
            for field, kws in PATTERNS:
                for kw in kws:
                    if kw in val:
                        if field not in col_map:
                            col_map[field] = c
                        break
        # 扫描元数据行（表头上方的单值单元格）
        # 元数据区模式：标签在左列，值在右列 → 匹配标签后偏移到值列
        meta_cells = {}
        for r in range(1, best_row):
            for c in range(1, max_col):
                val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                if not val:
                    continue
                for field, kws in PATTERNS:
                    for kw in kws:
                        # 元数据标签需精确匹配，避免「渠道」误匹配「渠道能力」
                        if val == kw and field not in col_map:
                            # 标签单元格 → 值在右边一列
                            meta_cells[field] = (r, c + 1)
                            break
        return best_row, col_map, meta_cells

    @staticmethod
    def _set_col(ws, row, col_idx, value, align=None, font=None, fill=None):
        c = ws.cell(row=row, column=col_idx)
        if isinstance(c, MergedCell):
            return
        c.value = value
        if align:
            c.alignment = align
        if font:
            c.font = font
        if fill:
            c.fill = fill

    def _extract_cell_images(self, wb_path: str) -> dict:
        """{row: (dispimg_name, img_bytes, w_emu, h_emu)}"""
        import zipfile as _zf, xml.etree.ElementTree as _ET, re as _re
        result = {}
        try:
            zf = _zf.ZipFile(wb_path)
            if "xl/cellimages.xml" not in zf.namelist():
                zf.close(); return result
            ci_xml = zf.read("xl/cellimages.xml")
            rels_xml = zf.read("xl/_rels/cellimages.xml.rels")
        except Exception:
            return result
        try:
            rels_root = _ET.fromstring(rels_xml)
            rid_to_media = {rel.get("Id"): rel.get("Target") for rel in rels_root}
            ns_xdr = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
            ns_a = "http://schemas.openxmlformats.org/drawingml/2006/main"
            ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
            ci_root = _ET.fromstring(ci_xml)
            name_info = {}
            for ci in ci_root:
                cNvPr = ci.find(f'.//{{{ns_xdr}}}cNvPr')
                if cNvPr is None: continue
                name = cNvPr.get("name", "")
                blip = ci.find(f'.//{{{ns_a}}}blip')
                rid = blip.get(f'{{{ns_r}}}embed') if blip is not None else None
                ext = ci.find(f'.//{{{ns_a}}}ext')
                cx = int(ext.get("cx")) if ext is not None else 0
                cy = int(ext.get("cy")) if ext is not None else 0
                name_info[name] = {"rid": rid, "cx": cx, "cy": cy}
            sheet_xml = zf.read("xl/worksheets/sheet1.xml")
            sheet_root = _ET.fromstring(sheet_xml)
            ns_s = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            for cell in sheet_root.iter(f"{{{ns_s}}}c"):
                ref = cell.get("r", "")
                m = _re.match(r"[A-Z]+(\d+)", ref)
                if not m: continue
                row = int(m.group(1))
                for f in cell.iter(f"{{{ns_s}}}f"):
                    if f.text and "DISPIMG" in f.text:
                        img_match = _re.search(r'"([^"]+)"', f.text)
                        if img_match:
                            ni = name_info.get(img_match.group(1))
                            if ni and ni["rid"] in rid_to_media:
                                img_bytes = zf.read("xl/" + rid_to_media[ni["rid"]])
                                result[row] = (img_match.group(1), img_bytes, ni["cx"], ni["cy"])
        finally:
            zf.close()
        return result

    def _inject_cellimages(self, output_path: str):
        import zipfile as _zf, os as _os, shutil as _sh
        tmp_path = output_path + ".tmp"
        WPS_CELLIMAGE_TYPE = "http://www.wps.cn/officeDocument/2020/cellImage"
        try:
            with _zf.ZipFile(output_path, 'r') as zin,                  _zf.ZipFile(tmp_path, 'w', _zf.ZIP_DEFLATED) as zout:
                # Find next available rId from workbook.xml.rels
                wb_rels = zin.read("xl/_rels/workbook.xml.rels").decode()
                max_rid = 0
                import re as _re
                for m in _re.finditer(r'Id="rId(\d+)"', wb_rels):
                    max_rid = max(max_rid, int(m.group(1)))
                next_rid = max_rid + 1
                wb_rels_patched = wb_rels.replace(
                    "</Relationships>",
                    f'<Relationship Id="rId{next_rid}" Type="{WPS_CELLIMAGE_TYPE}" '
                    f'Target="cellimages.xml"/></Relationships>')
                for item in zin.infolist():
                    if item.filename == "[Content_Types].xml":
                        ct_xml = zin.read(item.filename).decode()
                        if "/xl/cellimages.xml" not in ct_xml:
                            ct_xml = ct_xml.replace(
                                "</Types>",
                                '<Override PartName="/xl/cellimages.xml" '
                                'ContentType="application/vnd.ms-excel.cellimages+xml"/></Types>')
                        zout.writestr(item, ct_xml)
                    elif item.filename == "xl/_rels/workbook.xml.rels":
                        zout.writestr(item, wb_rels_patched)
                    else:
                        zout.writestr(item, zin.read(item.filename))
                # Inject cellimages + media from template
                with _zf.ZipFile(str(self._template_path), 'r') as tpl:
                    for name in tpl.namelist():
                        if 'cellimage' in name.lower():
                            zout.writestr(name, tpl.read(name))
                    for name in tpl.namelist():
                        if name.startswith('xl/media/') and tpl.getinfo(name).file_size > 0:
                            if name not in {i.filename for i in zin.infolist()}:
                                zout.writestr(name, tpl.read(name))
            _sh.move(tmp_path, output_path)
        except Exception:
            if _os.path.exists(tmp_path):
                _os.unlink(tmp_path)