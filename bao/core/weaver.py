"""FBA 装箱单编织引擎

将 FBA 原始货件数据转换为物流商模板行格式。
"""
from pathlib import Path
from typing import Optional

import openpyxl


# HS 编码参考表路径
_HS_PATH = Path(__file__).parent.parent.parent / "data" / "hs_code.xlsx"

# 默认 HS 编码（蜡烛类）
_DEFAULT_HS_CODE = "3406000090"


def load_hs_map() -> dict:
    hs_map = {}; hs_import = {}; hs_code_map = {}; fields = {}
    if not _HS_PATH.exists():
        return {"by_keyword": hs_map, "by_import": hs_import, "by_hs": hs_code_map, "fields": fields}
    wb = openpyxl.load_workbook(_HS_PATH, data_only=True)
    ws = wb.active
    col_names = [str(ws.cell(row=1,column=c).value or "").strip() for c in range(1,ws.max_column+1)]
    for r in range(2, ws.max_row+1):
        row_vals = [str(ws.cell(row=r,column=c).value or "").strip() for c in range(1,ws.max_column+1)]
        kw = row_vals[0] if row_vals else ""; ic = row_vals[1] if len(row_vals)>1 else ""; hc = row_vals[2] if len(row_vals)>2 else ic
        if kw and hc: hs_map[kw] = hc
        if ic and hc: hs_import[ic] = hc; hs_code_map[ic] = ic
        if hc:
            hs_code_map[hc] = hc
            info = {col_names[i]: row_vals[i] for i in range(min(len(col_names),len(row_vals))) if row_vals[i]}
            fields[hc] = info
            if kw:
                fields[kw] = info
            # 备注拆成短关键词（滑动窗口2-3字）
            bz = row_vals[7] if len(row_vals) > 7 else ""
            if bz:
                for win in [2, 3]:
                    for i in range(len(bz) - win + 1):
                        frag = bz[i:i+win]
                        if frag not in hs_map:
                            hs_map[frag] = hc
                            fields[frag] = info
                if bz not in hs_map:
                    hs_map[bz] = hc
                    fields[bz] = info
    wb.close()
    return {"by_keyword": hs_map, "by_import": hs_import, "by_hs": hs_code_map, "fields": fields}


def _box_range_normalize(box_range: str) -> str:
    """标准化箱号段：~ → -，单整数 → N-N"""
    if not box_range:
        return ""
    val = box_range.strip().replace("～", "-").replace("~", "-")
    # 纯整数（无分隔符）→ 补全为 N-N
    if val.isdigit():
        val = f"{val}-{val}"
    return val


def weave_fba(data: dict, hs_code_override: Optional[str] = None) -> dict:
    """将 FBA 解析结果编织为模板数据

    Args:
        data: FBAParser.parse() 返回的字典
        hs_code_override: 强制覆盖 HS 编码

    Returns:
        {
            "shipment_id": "FBA15LRB2LTZ",
            "total_boxes": 45,
            "total_weight": 816.9,
            "total_cbm": 3.191,
            "rows": [...]
        }
    """
    meta = data.get("meta", {})
    items = data.get("items", [])

    shipment_id = meta.get("shipment_id", "")
    total_boxes = meta.get("total_boxes", 0)

    hs_data = load_hs_map()
    by_keyword = hs_data["by_keyword"]
    by_import = hs_data["by_import"]
    by_hs = hs_data["by_hs"]
    fields = hs_data["fields"]
    rows = []
    total_weight = 0.0
    total_cbm = 0.0

    for item in items:
        box_range_raw = item.get("box_range", "")
        box_range = _box_range_normalize(box_range_raw)
        box_count = item.get("box_count", 0)
        weight = item.get("weight", 0)
        length = item.get("length", 0)
        width = item.get("width", 0)
        height = item.get("height", 0)
        declared_qty = int(item.get("declared_qty", 0))

        # HS 编码匹配：优先进口海关编码 → HS CODE → 关键词
        input_hs = item.get("input_hs_code", "")
        matched_kw = None
        hs_code = hs_code_override or _DEFAULT_HS_CODE

        if not hs_code_override:
            if input_hs and input_hs in by_import:
                hs_code = by_import[input_hs]
            elif input_hs and input_hs in by_hs:
                hs_code = by_hs.get(input_hs, hs_code)
            else:
                product_name = item.get("product_name") or item.get("msku", "")
                # 货件名称辅助匹配
                shipment_name = meta.get("shipment_name", "")
                match_text = shipment_name + " " + product_name if shipment_name else product_name
                matched_kw = None
                for keyword, code in by_keyword.items():
                    if keyword in match_text:
                        hs_code = code
                        matched_kw = keyword
                        break

        # 获取 HS 表完整字段信息（优先按关键词匹配，回退按HS）
        hs_fields = fields.get(matched_kw) or fields.get(hs_code, {})

        rows.append({
            "箱号段": box_range,
            "总件数": box_count,
            "SKU": item.get("msku", ""),
            "ASIN": item.get("asin", ""),
            "标题": item.get("title", ""),
            "进口海关编码": item.get("input_hs_code"),
            "海关编码": hs_code,
            "英文品名": hs_fields.get("英文品名", ""),
            "中文品名": hs_fields.get("中文品名", ""),
            "品牌": hs_fields.get("备注", ""),
            "amazon_title": item.get("title", ""),
            "材质": hs_fields.get("材质", ""),
            "用途": hs_fields.get("用途", ""),
            "总数量": declared_qty,
            "单箱重量": weight,
            "长": length,
            "宽": width,
            "高": height,
        })

        total_weight += weight * box_count
        # 先累加未舍入的体积，最后统一 round
        total_cbm += length * width * height * box_count / 1_000_000

    if total_boxes <= 0:
        total_boxes = sum(r["总件数"] for r in rows)

    return {
        "shipment_id": shipment_id,
        "shipment_name": meta.get("shipment_name", ""),
        "country_code": meta.get("country_code", ""),
        "total_boxes": total_boxes,
        "total_weight": round(total_weight, 2),
        "total_cbm": round(total_cbm, 3),
        "rows": rows,
    }
