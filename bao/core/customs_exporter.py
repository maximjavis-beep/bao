"""报关单信息汇总 Excel 导出器"""
import io
from typing import Dict, List
import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill


HEADERS = [
    "单号", "下单日期", "报单订单单号", "报关单号", "报关金额",
    "币种", "兑成RMB", "报关单是否收到", "平台", "上传政府平台",
    "申报日期", "提单报关单上传一达通平台", "物流", "报关抬头",
]

HEADER_FILL = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def export_customs_summary(items: List[Dict], output_path: str = None) -> bytes:
    """将报关单解析结果导出为 Excel

    Args:
        items: 解析后的报关单条目列表
        output_path: 可选，保存到文件路径；不传则返回 bytes

    Returns:
        Excel 文件的 bytes（output_path 为 None 时）
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报关单信息汇总"

    # ── 写表头 ──
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # ── 写数据行 ──
    for row_idx, item in enumerate(items, start=2):
        if "error" in item:
            # 出错的条目标记在 A 列
            ws.cell(row=row_idx, column=1, value=f"❌ {item.get('file','')}: {item['error']}")
            continue
        for col_idx, header in enumerate(HEADERS, start=1):
            val = item.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER

    # ── 列宽自适应 ──
    col_widths = {
        1: 22,   # 单号
        2: 14,   # 下单日期
        3: 14,   # 报单订单单号
        4: 22,   # 报关单号
        5: 14,   # 报关金额
        6: 10,   # 币种
        7: 14,   # 兑成RMB
        8: 16,   # 报关单是否收到
        9: 12,   # 平台
        10: 16,  # 上传政府平台
        11: 14,  # 申报日期
        12: 24,  # 提单报关单上传一达通平台
        13: 10,  # 物流
        14: 12,  # 报关抬头
    }
    for col, width in col_widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # ── 冻结表头 ──
    ws.freeze_panes = "A2"

    # ── 输出 ──
    if output_path:
        wb.save(output_path)
        wb.close()
        return None
    else:
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        return buf.getvalue()
