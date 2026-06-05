"""bao Web API — FastAPI for Vercel Serverless (secure: uploads deleted after use, downloads via one-shot tokens)"""
import base64, io, os, shutil, sys, uuid, zipfile
from pathlib import Path
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from fastapi import FastAPI, UploadFile, File, Form, Request
from fastapi.responses import JSONResponse, HTMLResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from bao.core.exporter import FBAExporter
from bao.core.customs_exporter import export_customs_summary
from bao.core.weaver import weave_fba
import openpyxl
from openpyxl.styles import PatternFill
from bao.parsers.customs_pdf_parser import CustomsPDFParser
from bao.parsers.fba_parser import FBAParser

app = FastAPI(title="bao", version="0.6.5")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])
ROOT_DIR = Path(__file__).parent.parent
UPLOAD_DIR = Path("/tmp/bao_uploads")
_DOWNLOAD_SLOTS = {}

BUILTIN_TEMPLATES = {
    "desu": ("德速-模板", "德速-模板.xlsx"),
    "jiufang": ("九方-模板", "九方-模版.xlsx"),
}
TEMPLATES_DIR = ROOT_DIR / "templates"

def _read_html(path): fp = ROOT_DIR / "bao" / "web" / path; return fp.read_text(encoding="utf-8") if fp.exists() else "<h1>404</h1>"

def _parse_tracking(path):
    """解析货件追踪码文件 → {FBA编号: {tracking_code, warehouse, channel, total_boxes}}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    col_map = {}
    for c in range(1, min(20, ws.max_column or 0) + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip()
        hl = h.lower()
        if "fba编号" in hl or "fba" in hl:
            col_map["fba"] = c
        elif "货件追踪" in h or "追踪码" in h or "tracking" in hl:
            col_map["tracking"] = c
        elif "仓库代码" in h or "仓库" in h:
            col_map["warehouse"] = c
        elif "渠道" in h:
            col_map["channel"] = c
        elif "箱数" in h or "总件数" in h:
            col_map["boxes"] = c
    if "fba" not in col_map:
        col_map["fba"] = 1
    if "tracking" not in col_map:
        col_map["tracking"] = 2
    result = {}
    for r in range(2, ws.max_row + 1):
        fba_id = str(ws.cell(row=r, column=col_map["fba"]).value or "").strip()
        if not fba_id:
            continue
        info = {"tracking_code": str(ws.cell(row=r, column=col_map["tracking"]).value or "").strip()}
        if "warehouse" in col_map:
            info["warehouse"] = str(ws.cell(row=r, column=col_map["warehouse"]).value or "").strip()
        if "channel" in col_map:
            info["channel"] = str(ws.cell(row=r, column=col_map["channel"]).value or "").strip()
        if "boxes" in col_map:
            v = ws.cell(row=r, column=col_map["boxes"]).value
            if v is not None:
                info["total_boxes"] = int(v) if isinstance(v, (int, float)) else str(v).strip()
        result[fba_id] = info
    wb.close()
    return result
def _parse_shipping(path):
    """出运数据 → {FBA编号: {warehouse, channel, total_boxes, tracking_code}}"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    result = {}
    col_map = {}
    for c in range(1, min(20, ws.max_column or 0) + 1):
        h = str(ws.cell(row=1, column=c).value or "").strip()
        hl = h.lower()
        if "fba编号" in hl or "fba" in hl:
            col_map["fba"] = c
        elif "货件追踪" in h or "追踪码" in h or "tracking" in hl:
            col_map["tracking"] = c
        elif "仓库代码" in h or "仓库" in h:
            col_map["warehouse"] = c
        elif "渠道" in h:
            col_map["channel"] = c
        elif "箱数" in h or "总件数" in h:
            col_map["boxes"] = c
    if "fba" not in col_map:
        col_map["fba"] = 2
    for r in range(2, ws.max_row + 1):
        fba_id = str(ws.cell(row=r, column=col_map.get("fba", 2)).value or "").strip()
        if not fba_id:
            continue
        info = {}
        if "warehouse" in col_map:
            info["warehouse"] = str(ws.cell(row=r, column=col_map["warehouse"]).value or "").strip()
        if "channel" in col_map:
            info["channel"] = str(ws.cell(row=r, column=col_map["channel"]).value or "").strip()
        if "boxes" in col_map:
            v = ws.cell(row=r, column=col_map["boxes"]).value
            if v is not None:
                info["total_boxes"] = int(v) if isinstance(v, (int, float)) else str(v).strip()
        if "tracking" in col_map:
            info["tracking_code"] = str(ws.cell(row=r, column=col_map["tracking"]).value or "").strip()
        result[fba_id] = info
    wb.close()
    return result
def _make_token(fp):
    if not fp.exists(): return ""
    b64 = base64.b64encode(fp.read_bytes()).decode()
    token = uuid.uuid4().hex; _DOWNLOAD_SLOTS[token] = b64
    fp.unlink(missing_ok=True); return token

def _resolve_template(template_id):
    if template_id and template_id in BUILTIN_TEMPLATES:
        tpl_file = TEMPLATES_DIR / BUILTIN_TEMPLATES[template_id][1]
        if tpl_file.exists():
            return str(tpl_file)
    return None

@app.get("/api/templates")
async def api_templates():
    tpls = [{"id": k, "name": v[0]} for k, v in BUILTIN_TEMPLATES.items()]
    return JSONResponse({"success": True, "templates": tpls})

@app.get("/", response_class=HTMLResponse)
async def index(): return _read_html("index_vercel.html")

@app.post("/api/parse")
async def api_parse(request: Request, file: UploadFile = File(None)):
    try:
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        content_type = request.headers.get("content-type", "")
        if "application/json" in content_type:
            body = await request.json()
            b64 = body.get("file", "")
            if not b64:
                return JSONResponse({"success": False, "error": "未提供文件"}, 400)
            tmp = UPLOAD_DIR / f"fba_{uuid.uuid4().hex[:6]}.xlsx"
            tmp.write_bytes(base64.b64decode(b64))
        elif file:
            tmp = UPLOAD_DIR / f"fba_{uuid.uuid4().hex[:6]}.xlsx"
            tmp.write_bytes(await file.read())
        else:
            return JSONResponse({"success": False, "error": "未提供文件"}, 400)
        data = FBAParser().parse(str(tmp)); tmp.unlink(missing_ok=True)
        items = data.get("items", [])
        return JSONResponse({"success": True, "meta": data.get("meta", {}), "items": items, "item_count": len(items)})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, 400)

@app.post("/api/weave")
async def api_weave(request: Request, file: UploadFile = File(None), hs_code: str = Form(None), template: UploadFile = File(None), tracking: UploadFile = File(None), shipping: UploadFile = File(None), template_id: str = Form(None)):
    try:
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        content_type = request.headers.get("content-type", "")
        if "application/json" in content_type:
            body = await request.json()
            b64 = body.get("file", "")
            if not b64: return JSONResponse({"success": False, "error": "未提供文件"}, 400)
            tmp = UPLOAD_DIR / f"fba_{uuid.uuid4().hex[:6]}.xlsx"
            tmp.write_bytes(base64.b64decode(b64))
            hs_code = body.get("hs_code") or hs_code
            tpl_b64 = body.get("template")
            track_b64 = body.get("tracking")
            ship_b64 = body.get("shipping")
            template_id = body.get("template_id") or template_id
        elif file:
            tmp = UPLOAD_DIR / f"fba_{uuid.uuid4().hex[:6]}.xlsx"
            tmp.write_bytes(await file.read())
        else:
            return JSONResponse({"success": False, "error": "未提供文件"}, 400)
        data = FBAParser().parse(str(tmp)); tmp.unlink(missing_ok=True)
        meta = data.get("meta", {})
        woven = weave_fba(data, hs_code_override=hs_code or None)
        tpl_path = None
        if "application/json" in content_type and locals().get('tpl_b64'):
            tpl_tmp = UPLOAD_DIR / f"tpl_{uuid.uuid4().hex[:6]}.xlsx"
            tpl_tmp.write_bytes(base64.b64decode(tpl_b64))
            tpl_path = str(tpl_tmp)
        elif template:
            tpl_tmp = UPLOAD_DIR / f"tpl_{uuid.uuid4().hex[:6]}.xlsx"
            tpl_tmp.write_bytes(await template.read())
            tpl_path = str(tpl_tmp)
        else:
            tpl_path = _resolve_template(template_id)
        tracking_map = None
        if "application/json" in content_type and locals().get('track_b64'):
            tk_tmp = UPLOAD_DIR / f"track_{uuid.uuid4().hex[:6]}.xlsx"
            tk_tmp.write_bytes(base64.b64decode(track_b64))
            tracking_map = _parse_tracking(str(tk_tmp))
        elif tracking:
            tk_tmp = UPLOAD_DIR / f"track_{uuid.uuid4().hex[:6]}.xlsx"
            tk_tmp.write_bytes(await tracking.read())
            tracking_map = _parse_tracking(str(tk_tmp))
        shipping_map = None
        if "application/json" in content_type and locals().get('ship_b64'):
            sh_tmp = UPLOAD_DIR / f"ship_{uuid.uuid4().hex[:6]}.xlsx"
            sh_tmp.write_bytes(base64.b64decode(ship_b64))
            shipping_map = _parse_shipping(str(sh_tmp))
        elif shipping:
            sh_tmp = UPLOAD_DIR / f"ship_{uuid.uuid4().hex[:6]}.xlsx"
            sh_tmp.write_bytes(await shipping.read())
            shipping_map = _parse_shipping(str(sh_tmp))
        out = UPLOAD_DIR / f"{meta.get('shipment_id','UNKNOWN')}-装箱单_{uuid.uuid4().hex[:6]}.xlsx"
        FBAExporter(template_path=tpl_path, tracking_map=tracking_map, shipping_map=shipping_map).export(woven, str(out))
        file_b64 = base64.b64encode(out.read_bytes()).decode()
        download_url = ""
        if out.exists():
            download_url = "/api/download/" + _make_token(out)
        else:
            out.unlink(missing_ok=True)
        return JSONResponse({"success": True, "file_b64": file_b64, "download_url": download_url, "summary": {
            "shipment_id": woven["shipment_id"], "total_boxes": woven["total_boxes"],
            "total_weight": woven["total_weight"], "total_cbm": woven["total_cbm"],
            "row_count": len(woven["rows"])}, "rows": woven["rows"]})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, 400)

@app.post("/api/weave-batch")
async def api_weave_batch(files: list[UploadFile] = File(...), hs_code: str = Form(None), template: UploadFile = File(None), tracking: UploadFile = File(None), shipping: UploadFile = File(None), template_id: str = Form(None)):
    try:
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        tpl_path = None
        if template:
            tpl_tmp = UPLOAD_DIR / f"tpl_{uuid.uuid4().hex[:6]}.xlsx"
            tpl_tmp.write_bytes(await template.read())
            tpl_path = str(tpl_tmp)
        else:
            tpl_path = _resolve_template(template_id)
        tracking_map = None
        if tracking:
            tk_tmp = UPLOAD_DIR / f"track_{uuid.uuid4().hex[:6]}.xlsx"
            tk_tmp.write_bytes(await tracking.read())
            tracking_map = _parse_tracking(str(tk_tmp))
        shipping_map = None
        if shipping:
            sh_tmp = UPLOAD_DIR / f"ship_{uuid.uuid4().hex[:6]}.xlsx"
            sh_tmp.write_bytes(await shipping.read())
            shipping_map = _parse_shipping(str(sh_tmp))
        parser, exporter, results = FBAParser(), FBAExporter(template_path=tpl_path, tracking_map=tracking_map, shipping_map=shipping_map), []
        for f in files:
            try:
                tmp = UPLOAD_DIR / f"fba_{uuid.uuid4().hex[:6]}.xlsx"
                tmp.write_bytes(await f.read())
                d = parser.parse(str(tmp)); tmp.unlink(missing_ok=True)
                meta = d.get("meta", {})
                woven = weave_fba(d, hs_code_override=hs_code or None)
                out = UPLOAD_DIR / f"{meta.get('shipment_id','UNKNOWN')}-装箱单_{uuid.uuid4().hex[:6]}.xlsx"
                exporter.export(woven, str(out))
                file_b64 = base64.b64encode(out.read_bytes()).decode()
                out.unlink(missing_ok=True)
                results.append({"filename": f.filename or "未知", "success": True,
                    "shipment_id": meta.get("shipment_id",""), "file_b64": file_b64,
                    "summary": {"shipment_id": woven["shipment_id"], "total_boxes": woven["total_boxes"],
                    "total_weight": woven["total_weight"], "total_cbm": woven["total_cbm"], "row_count": len(woven["rows"])}})
            except Exception as e:
                results.append({"filename": f.filename or "未知", "success": False, "error": str(e)})
        zip_b64 = ""
        zip_error = ""
        ok = [r for r in results if r.get("success")]
        if ok:
            try:
                bid = uuid.uuid4().hex[:6]
                td = UPLOAD_DIR / f"batch_{bid}"
                td.mkdir(parents=True, exist_ok=True)
                for r in ok:
                    fb = r.get("file_b64","")
                    if fb:
                        (td / f"{r['shipment_id']}.xlsx").write_bytes(base64.b64decode(fb))
                # 直接用 Python zipfile，跳过 subprocess（Vercel 不保证有 zip 命令）
                zd = str(UPLOAD_DIR / f"batch_{bid}.zip")
                zf = zipfile.ZipFile(zd, "w", zipfile.ZIP_DEFLATED)
                for fl in sorted(td.iterdir()):
                    zf.write(str(fl), fl.name)
                zf.close()
                zip_b64 = base64.b64encode(Path(zd).read_bytes()).decode()
            except Exception as e:
                zip_error = str(e)[:200]
            finally:
                try:
                    shutil.rmtree(td, ignore_errors=True)
                    Path(zd).unlink(missing_ok=True)
                except Exception:
                    pass
        return JSONResponse({"success": True, "results": results, "zip_b64": zip_b64, "zip_error": zip_error, "result_count": len(results)})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, 400)

@app.post("/api/merge-plan")
async def api_merge_plan(file: UploadFile = File(...), plan_file: UploadFile = File(None)):
    """上传调整明细 + 发货计划，按四步规则调整后输出"""
    try:
        if not plan_file:
            return JSONResponse({"success": False, "error": "请同时上传调整明细和发货计划"}, 400)
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)

        tmp_adj = UPLOAD_DIR / f"adj_{uuid.uuid4().hex[:6]}.xlsx"
        tmp_adj.write_bytes(await file.read())
        tmp_plan = UPLOAD_DIR / f"plan_{uuid.uuid4().hex[:6]}.xlsx"
        tmp_plan.write_bytes(await plan_file.read())

        # 解析调整明细
        wb_adj = openpyxl.load_workbook(str(tmp_adj), data_only=True)
        ws_adj = wb_adj.active
        col_idx = {}
        for c in range(1, ws_adj.max_column + 1):
            h = str(ws_adj.cell(row=1, column=c).value or "")
            if h.strip() == "SKU": col_idx["SKU"] = c
            elif "调整单号" in h: col_idx["调整单号"] = c
            elif "识别码" in h and "新" not in h: col_idx["识别码"] = c
            elif "调整FNSKU" in h: col_idx["调整FNSKU"] = c
            elif "调整量" in h: col_idx["调整量"] = c
            elif "调整店铺" in h: col_idx["调整店铺"] = c
            elif "店铺" in h and "调整" not in h: col_idx["店铺"] = c
        if "SKU" not in col_idx or "识别码" not in col_idx:
            wb_adj.close(); tmp_adj.unlink(missing_ok=True); tmp_plan.unlink(missing_ok=True)
            return JSONResponse({"success": False, "error": "未找到SKU或识别码列"}, 400)
        adjustments = []
        for r in range(2, ws_adj.max_row + 1):
            code = ws_adj.cell(row=r, column=col_idx["识别码"]).value
            if code is None: continue
            qty = ws_adj.cell(row=r, column=col_idx.get("调整量",0)).value
            adjustments.append({
                "调整单号": str(ws_adj.cell(row=r, column=col_idx.get("调整单号",0)).value or "").strip(),
                "识别码": str(code).strip(),
                "店铺": str(ws_adj.cell(row=r, column=col_idx.get("店铺",0)).value or "").strip(),
                "SKU": str(ws_adj.cell(row=r, column=col_idx["SKU"]).value or "").strip(),
                "调整FNSKU": str(ws_adj.cell(row=r, column=col_idx.get("调整FNSKU",0)).value or "").strip(),
                "调整量": int(qty) if qty is not None else 0,
                "调整店铺": str(ws_adj.cell(row=r, column=col_idx.get("调整店铺",0)).value or "").strip(),
            })
        wb_adj.close(); tmp_adj.unlink(missing_ok=True)
        if not adjustments:
            tmp_plan.unlink(missing_ok=True)
            return JSONResponse({"success": False, "error": "调整明细无有效数据"}, 400)
        wb_plan = openpyxl.load_workbook(str(tmp_plan))
        ws_plan = wb_plan.active
        plan_rows = []
        for r in range(2, ws_plan.max_row + 1):
            plan_rows.append({
                "r": r,
                "识别码": str(ws_plan.cell(row=r, column=3).value or "").strip(),
                "FNSKU": str(ws_plan.cell(row=r, column=6).value or "").strip(),
                "店铺": str(ws_plan.cell(row=r, column=15).value or "").strip(),
                "计划发货量": int(ws_plan.cell(row=r, column=9).value) if ws_plan.cell(row=r, column=9).value else 0,
                "箱数": int(ws_plan.cell(row=r, column=8).value) if ws_plan.cell(row=r, column=8).value else 0,
                "is_new": (ws_plan.cell(row=r, column=1).value is None),
            })
        match_map = {}
        for adj in adjustments:
            for pr in plan_rows:
                if adj["调整店铺"] == pr["店铺"] and adj["识别码"] == pr["识别码"] and adj["调整FNSKU"] == pr["FNSKU"]:
                    key = (adj["识别码"], adj["调整FNSKU"])
                    if key not in match_map:
                        match_map[key] = {"plan_row": pr, "adjustments": [], "total_adj_qty": 0}
                    match_map[key]["adjustments"].append(adj)
                    match_map[key]["total_adj_qty"] += adj["调整量"]
                    break
        deleted = set()
        for key, data in match_map.items():
            pr = data["plan_row"]
            adj_list = data["adjustments"]
            r = pr["r"]
            adj_nos = ",".join(sorted(set(a["调整单号"] for a in adj_list)))
            ws_plan.cell(row=r, column=17).value = adj_nos
            ws_plan.cell(row=r, column=4).value = adj_list[0]["调整店铺"]
            v = ws_plan.cell(row=r, column=11).value
            if v is None or (isinstance(v, str) and v.startswith("=")):
                ws_plan.cell(row=r, column=11).value = pr["箱数"]
        for pr in plan_rows:
            if pr["is_new"]:
                key = (pr["识别码"], pr["FNSKU"])
                if key in match_map:
                    for c in range(1, ws_plan.max_column + 1):
                        ws_plan.cell(row=pr["r"], column=c).value = None
                    deleted.add(pr["r"])
        for pr in plan_rows:
            if pr["r"] not in deleted and pr["箱数"] > 0:
                v = ws_plan.cell(row=pr["r"], column=11).value
                if v is None or (isinstance(v, str) and v.startswith("=")):
                    ws_plan.cell(row=pr["r"], column=11).value = pr["箱数"]
        out = UPLOAD_DIR / f"发货计划_{uuid.uuid4().hex[:6]}.xlsx"
        wb_plan.save(str(out)); wb_plan.close(); tmp_plan.unlink(missing_ok=True)
        file_b64 = base64.b64encode(out.read_bytes()).decode()
        out.unlink(missing_ok=True)
        return JSONResponse({"success": True, "file_b64": file_b64, "matched": len(match_map), "deleted": len(deleted)})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, 400)

@app.post("/api/customs-summary")
async def api_customs_summary(request: Request):
    """报关单 PDF 解析并生成汇总 Excel"""
    try:
        body = await request.json()
        files_data = body.get("files", [])
        if not files_data:
            return JSONResponse({"success": False, "error": "未上传任何 PDF 文件"}, 400)
        parser = CustomsPDFParser()
        items = []
        UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        for f in files_data:
            try:
                b64 = f.get("data", "")
                name = f.get("name", "unknown.pdf")
                if not b64:
                    items.append({"error": "文件数据为空", "file": name})
                    continue
                pdf_path = UPLOAD_DIR / f"customs_{uuid.uuid4().hex[:6]}.pdf"
                pdf_path.write_bytes(base64.b64decode(b64))
                result = parser.parse(str(pdf_path))
                result["_source"] = name
                items.append(result)
                pdf_path.unlink(missing_ok=True)
            except Exception as e:
                items.append({"error": str(e), "file": f.get("name", "unknown.pdf")})
        file_b64 = ""
        if any("error" not in it for it in items):
            xlsx_bytes = export_customs_summary(items)
            file_b64 = base64.b64encode(xlsx_bytes).decode()
        ok_items = [it for it in items if "error" not in it]
        total_amount = sum(it.get("报关金额", 0) or 0 for it in ok_items if isinstance(it.get("报关金额"), (int, float)))
        currencies = list(set(it.get("币种", "") for it in ok_items if it.get("币种")))
        return JSONResponse({"success": True, "items": items, "item_count": len(items), "ok_count": len(ok_items), "file_b64": file_b64, "stats": {"total": len(items), "ok": len(ok_items), "total_amount": round(total_amount, 2), "currencies": currencies}})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, 400)

@app.post("/api/customs-export")
async def api_customs_export(request: Request):
    """根据已解析的报关单数据生成汇总 Excel"""
    try:
        body = await request.json()
        items = body.get("items", [])
        if not items:
            return JSONResponse({"success": False, "error": "无数据可导出"}, 400)
        xlsx_bytes = export_customs_summary(items)
        file_b64 = base64.b64encode(xlsx_bytes).decode()
        return JSONResponse({"success": True, "file_b64": file_b64})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, 400)

@app.get("/api/download/{token}")
async def dl_token(token: str):
    b64 = _DOWNLOAD_SLOTS.pop(token, None)
    if not b64: return Response("Not Found or expired", status_code=404)
    return Response(content=base64.b64decode(b64), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=装箱单.xlsx"})

@app.post("/api/download-zip")
async def dl_zip(request: Request):
    """接收 [{name, b64}] 打包成 ZIP 返回"""
    try:
        body = await request.json()
        files = body.get("files", [])
        if not files:
            return JSONResponse({"success": False, "error": "无文件"}, 400)
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in files:
                zf.writestr(f.get("name", "unknown.xlsx"), base64.b64decode(f.get("b64", "")))
        zip_b64 = base64.b64encode(buf.getvalue()).decode()
        return JSONResponse({"success": True, "zip_b64": zip_b64})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, 400)