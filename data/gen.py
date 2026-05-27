"""生成示例发票和装箱单 — 标准格式"""
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

OUT = "/Users/streiten/customs/bao/data"
thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
b = Font(bold=True)

# ============ 发票 ============
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "商业发票"

# 表头信息: label 在 A 列，value 在 B 列
hdr = [
    ("发票号", "INV-2026-0521"),
    ("发票日期", "2026-05-21"),
    ("卖方", "苏州恒通纺织进出口有限公司"),
    ("买方", "SUNSHINE TEXTILE TRADING LTD."),
    ("合同号", "SC-2026-0415"),
    ("贸易术语", "FOB"),
    ("币制", "USD"),
    ("起运港", "上海"),
    ("目的港", "洛杉矶"),
    ("船名", "EVER FORWARD"),
    ("件数", "120"),
    ("包装", "纸箱"),
]
for i, (k, v) in enumerate(hdr, 1):
    ws.cell(i, 1, k).font = b
    ws.cell(i, 2, v)

# 空两行后放数据表
r = len(hdr) + 3
cols = ["序号","HS编码","品名","英文品名","规格型号","数量","单位","单价","总价","净重(KG)","毛重(KG)"]
for ci, h in enumerate(cols, 1):
    c = ws.cell(r, ci, h)
    c.font = b
    c.border = thin

items = [
    [1,"5208.4200","全棉色织布","Cotton Yarn-Dyed Fabric","40Sx40S 133x72 57/58\"",5000,"米",3.5,17500,1200,1280],
    [2,"5209.3200","全棉帆布","Cotton Canvas","10Sx10S 72x40 58\"",3000,"米",5.2,15600,1800,1900],
    [3,"5513.4100","涤棉府绸","TC Poplin Fabric","45Sx45S 110x76 58\"",4000,"米",2.8,11200,960,1040],
]
for item in items:
    r += 1
    for ci, v in enumerate(item, 1):
        ws.cell(r, ci, v).border = thin

r += 1
ws.cell(r, 1, "合计").font = b
ws.cell(r, 9, 44300).font = b

wb.save(f"{OUT}/示例发票.xlsx")
print("✅ 示例发票已生成")

# ============ 装箱单 ============
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "装箱单"

pkg_hdr = [
    ("发票号", "INV-2026-0521"),
    ("日期", "2026-05-21"),
    ("合同号", "SC-2026-0415"),
    ("发货人", "苏州恒通纺织进出口有限公司"),
    ("收货人", "SUNSHINE TEXTILE TRADING LTD."),
    ("包装", "纸箱 / 120 CTNS"),
    ("唛头", "SUNSHINE / L.A. / C/NO.1-120"),
]
for i, (k, v) in enumerate(pkg_hdr, 1):
    ws2.cell(i, 1, k).font = b
    ws2.cell(i, 2, v)

r = len(pkg_hdr) + 3
pkg_cols = ["箱号","品名","数量","净重(KG)","毛重(KG)","尺寸(cm)"]
for ci, h in enumerate(pkg_cols, 1):
    c = ws2.cell(r, ci, h)
    c.font = b
    c.border = thin

for i in range(1, 121):
    if i <= 50:
        d, q, nw, gw, sz = "全棉色织布", 100, 24, 25.6, "50x40x30"
    elif i <= 75:
        d, q, nw, gw, sz = "全棉帆布", 120, 72, 76, "50x40x35"
    else:
        d, q, nw, gw, sz = "涤棉府绸", 88.89, 21.33, 23.11, "50x40x30"
    r += 1
    for ci, v in enumerate([i, d, q, round(nw,1), round(gw,1), sz], 1):
        ws2.cell(r, ci, v).border = thin

r += 1
ws2.cell(r, 1, "合计").font = b
ws2.cell(r, 3, 12000).font = b

wb2.save(f"{OUT}/示例装箱单.xlsx")
print("✅ 示例装箱单已生成")
